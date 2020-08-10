#! python3
# coding: utf-8

#Excelの操作に
import openpyxl
import json

with open('Settings.json') as jf:
    settingFile = json.load(jf)
column = settingFile['column']
memberProfile = settingFile['profile']
cellRange = settingFile['range']


#ファイルを開く
# data_only=True 関数のセルも値を読んでくる
book = openpyxl.load_workbook('templete_cell.xlsx', data_only=True)
sheet = book['1']

#勤務時間入力
for i in range(cellRange['start'], cellRange['end']):
    day = sheet.cell(row=i, column=column['day']).value
    if day == '土' or day == '日' or day == None or day == '*':
        continue
    else:
        sheet.cell(i + 1, column['opening'], '9:00')
        sheet.cell(i + 1, column['closing'], '17:30')


with open('NameList.json') as n:
    nameFile = json.load(n)
year_month = str(nameFile['year_month'])
memberList = nameFile['member']

for member in memberList:
    #profile入力
    sheet[memberProfile['name']] = member['name']
    sheet[memberProfile['yaku']] = member['yaku']

    #保存
    book.save(year_month + '_' + member['name'] + '.xlsx')